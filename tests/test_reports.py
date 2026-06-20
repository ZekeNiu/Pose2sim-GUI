from __future__ import annotations

import json
from pathlib import Path
import re
import tempfile
import unittest

from openpyxl import load_workbook

from pose2sim_gui.reports import (
    coordinate_infos_for_columns,
    export_excel,
    export_html,
    find_report_video_files,
    read_mot,
    split_motion_columns,
)


MOT_TEXT = """Coordinates
version=1
nRows=3
nColumns=5
inDegrees=yes
endheader
time pelvis_tx hip_flexion_r knee_angle_r ankle_angle_r
0.00 0.10 12.0 30.0 5.0
0.01 0.11 13.0 31.0 6.0
0.02 0.12 14.0 32.0 7.0
"""


class ReportTests(unittest.TestCase):
    def test_read_mot_and_split_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            mot = Path(tmp) / "trial.mot"
            mot.write_text(MOT_TEXT, encoding="utf-8")
            data, header = read_mot(mot)
            self.assertEqual(len(header), 6)
            self.assertEqual(list(data.columns), ["time", "pelvis_tx", "hip_flexion_r", "knee_angle_r", "ankle_angle_r"])
            angles, translations = split_motion_columns(data)
            self.assertEqual(translations, ["pelvis_tx"])
            self.assertIn("hip_flexion_r", angles)

    def test_export_excel(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            mot = Path(tmp) / "trial.mot"
            mot.write_text(MOT_TEXT, encoding="utf-8")
            output = export_excel(mot)
            self.assertTrue(output.exists())
            workbook = load_workbook(output, read_only=True)
            try:
                self.assertEqual(set(workbook.sheetnames), {"关节活动度", "平移坐标", "统计摘要"})
                self.assertEqual(workbook["关节活动度"]["A1"].value, "时间（秒）")
                self.assertEqual(workbook["关节活动度"]["B1"].value, "右髋关节屈伸活动度")
                self.assertEqual(workbook["平移坐标"]["B1"].value, "pelvis_tx")
                self.assertEqual(workbook["统计摘要"]["A1"].value, "指标")
            finally:
                workbook.close()

    def test_export_html_contains_plot_video_and_sync_script(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            mot = tmp_path / "trial.mot"
            video1 = tmp_path / "cam01_pose.mp4"
            video2 = tmp_path / "cam02_pose.mp4"
            mot.write_text(MOT_TEXT, encoding="utf-8")
            video1.write_bytes(b"")
            video2.write_bytes(b"")
            output, warnings = export_html(mot, [video1, video2], transcode_video=False)
            self.assertEqual(warnings, [])
            html = output.read_text(encoding="utf-8")
            self.assertIn("Plotly.newPlot", html)
            self.assertEqual(html.count("<video"), 2)
            self.assertIn("media/01_cam01_pose.mp4", html)
            self.assertIn('type="video/mp4"', html)
            self.assertIn("plotly_hover", html)
            self.assertIn("右髋关节屈伸活动度", html)
            self.assertIn("当前时刻关节活动度", html)
            self.assertIn("完整统计表", html)
            self.assertIn("查看完整诊断", html)
            self.assertIn('class="video-grid video-count-2"', html)
            self.assertNotIn("metric-toggle", html)
            self.assertNotIn("select-action", html)
            payload_match = re.search(r'<script id="motionData" type="application/json">(.*?)</script>', html, re.S)
            self.assertIsNotNone(payload_match)
            payload = json.loads(payload_match.group(1))
            self.assertEqual(payload["translationColumns"], ["pelvis_tx"])
            self.assertNotIn("pelvis_tx", payload["angleColumns"])
            self.assertEqual(payload["labels"]["knee_angle_r"], "右膝关节屈伸活动度")

    def test_coordinate_labels_are_complete_and_strict(self) -> None:
        infos = coordinate_infos_for_columns(
            ["hip_flexion_l", "hip_flexion_r", "knee_angle_l", "knee_angle_r", "flexion_l"]
        )
        labels = [info["label"] for info in infos.values()]
        self.assertIn("左髋关节屈伸活动度", labels)
        self.assertIn("右髋关节屈伸活动度", labels)
        self.assertIn("左膝关节屈伸活动度", labels)
        self.assertIn("右膝关节屈伸活动度", labels)
        self.assertEqual(infos["flexion_l"]["label"], "左未识别活动度 1")
        self.assertFalse(infos["flexion_l"]["recognized"])
        self.assertNotIn("屈伸活动度", infos["flexion_l"]["label"])

    def test_duplicate_labels_are_numbered(self) -> None:
        infos = coordinate_infos_for_columns(["knee_angle_l", "left_knee_angle"])
        self.assertEqual(infos["knee_angle_l"]["label"], "左膝关节屈伸活动度（1）")
        self.assertEqual(infos["left_knee_angle"]["label"], "左膝关节屈伸活动度（2）")

    def test_find_report_video_files_prefers_processed_videos(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp)
            pose = project / "pose"
            videos = project / "videos"
            pose.mkdir()
            videos.mkdir()
            processed = pose / "cam01_pose.mp4"
            raw = videos / "cam01.mp4"
            processed.write_bytes(b"")
            raw.write_bytes(b"")
            self.assertEqual(find_report_video_files(project), [processed.resolve()])


if __name__ == "__main__":
    unittest.main()
